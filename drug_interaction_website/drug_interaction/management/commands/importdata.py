from argparse import FileType
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook


from drug_interaction.models import Drug, Interaction


class Command(BaseCommand):
    help = 'Imports data from interactions Excel file into the database'
    requires_migrations_checks = True

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel_file', type=FileType('rb'), help='The Excel file containing the drug interaction data',
            default=settings.EXCEL_FILE_PATH
        )

    def handle(self, *args, **options):
        self.stdout.write(self.style.NOTICE(f'Importing data from {options["excel_file"].name} ...'))
        wb = load_workbook(options['excel_file'], read_only=True, data_only=True)
        ws = wb.active
        drugs = {}
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            drug1_name, drug2_name, reason = row
            if drug1_name not in drugs:
                drugs[drug1_name] = Drug.objects.get_or_create(name=drug1_name)[0]
            if drug2_name not in drugs:
                drugs[drug2_name] = Drug.objects.get_or_create(name=drug2_name)[0]
            Interaction.objects.get_or_create(drug1=drugs[drug1_name], drug2=drugs[drug2_name], reason=reason)
        self.stdout.write(self.style.SUCCESS('Successfully imported data'))
